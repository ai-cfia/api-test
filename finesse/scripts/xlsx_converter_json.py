import openpyxl
import os
import json
import argparse

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='XLSX Converter to JSON')
    parser.add_argument('--input-folder', dest='input_folder', default='finesse/scripts/input/', help='Path to the input folder')
    parser.add_argument('--output-folder', dest='output_folder', default='finesse/scripts/output/', help='Path to the output folder')
    parser.add_argument('--file-name', dest='file_name', help='Name of the input file')
    parser.add_argument('--sheet-name', dest='sheet_name', default='To fill', help='Name of the worksheet')

    args = parser.parse_args()

    INPUT_FOLDER = args.input_folder
    OUTPUT_FOLDER = args.output_folder
    FILE_NAME = args.file_name
    SHEET_NAME = args.sheet_name
    FILE_PATH = INPUT_FOLDER+FILE_NAME

    workbook = openpyxl.load_workbook(FILE_PATH)
    worksheet = workbook.active
    count = 1

    for row in range(5, worksheet.max_row + 1):
        question = worksheet.cell(row=row, column=2).value
        if question is None:
            continue

        answer = worksheet.cell(row=row, column=3).value

        titles = []
        links = []
        for col in range(5, 10):
            title = worksheet.cell(row=row, column=col).value
            link = worksheet.cell(row=row, column=col).hyperlink
            if title:
                titles.append(title)
            if link:
                links.append(link.target)

        data = {
            'question': question or "",
            'answer': answer or "",
            'title': titles[0] if len(titles) == 1 else titles or "",
            'url': links[0] if len(links) == 1 else links or ""
        }

        # Enregistrement du fichier JSON
        output_file = os.path.join(OUTPUT_FOLDER, f'question_{count}.json')
        if not os.path.exists(OUTPUT_FOLDER):
            os.makedirs(OUTPUT_FOLDER)
        with open(output_file, 'w', encoding='utf-8') as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=4)
            json_file.write('\n')
        count += 1

    print("Conversion completed successfully!")
